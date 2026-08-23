"""Parse account lists from CSV, XLSX, and JSON files.

Accepted formats:
  - CSV: first row may be a header; expected columns contain email and password
  - XLSX: first row may be a header; expected columns contain email and password
  - JSON: a list of objects with "email"/"password", or a list of [email, password]
"""
import csv
import io
import json
from typing import List, Tuple

EMAIL_KEYS = {"email", "mail", "username", "user", "login", "account", "e-mail", "email_address"}
PASSWORD_KEYS = {"password", "pass", "pw", "passwd", "pwd", "password1"}


def _pick_columns(header: List[str]) -> Tuple[int, int]:
    """Return (email_col, password_col) indexes from a header row."""
    email_col = password_col = None
    for i, cell in enumerate(header):
        key = str(cell).strip().lower()
        if email_col is None and key in EMAIL_KEYS:
            email_col = i
        if password_col is None and key in PASSWORD_KEYS:
            password_col = i
    return email_col, password_col


def parse_import(content: bytes, filename: str) -> List[dict]:
    """Parse uploaded file content into a list of {email, password} dicts."""
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt")):
        return _parse_csv(content)
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    if name.endswith(".json"):
        return _parse_json(content)
    raise ValueError("Unsupported file type. Please upload a CSV, XLSX, or JSON file.")


def _parse_csv(content: bytes) -> List[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader if row and any(str(c).strip() for c in row)]

    if not rows:
        return []

    email_col, password_col = _pick_columns(rows[0])
    start = 0
    if email_col is not None or password_col is not None:
        start = 1
    else:
        email_col, password_col = 0, 1

    accounts = []
    for row in rows[start:]:
        if len(row) <= max(email_col, password_col):
            continue
        email = str(row[email_col]).strip()
        password = str(row[password_col]).strip() if password_col < len(row) else ""
        if email and "@" in email:
            accounts.append({"email": email, "password": password})
    return accounts


def _parse_xlsx(content: bytes) -> List[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = [str(c).strip() if c is not None else "" for c in row]
        if any(values):
            rows.append(values)

    if not rows:
        return []

    email_col, password_col = _pick_columns(rows[0])
    start = 0
    if email_col is not None or password_col is not None:
        start = 1
    else:
        email_col, password_col = 0, 1

    accounts = []
    for row in rows[start:]:
        if len(row) <= max(email_col, password_col):
            continue
        email = row[email_col]
        password = row[password_col] if password_col < len(row) else ""
        if email and "@" in email:
            accounts.append({"email": email, "password": password})
    return accounts


def _parse_json(content: bytes) -> List[dict]:
    data = json.loads(content.decode("utf-8-sig", errors="replace"))
    accounts = []

    if isinstance(data, dict):
        # {"email": "password"}
        for email, password in data.items():
            accounts.append({"email": email, "password": str(password)})
        return accounts

    if not isinstance(data, list):
        raise ValueError("JSON must be an array of objects or a dict of email->password.")

    for item in data:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            email = str(item[0]).strip()
            password = str(item[1]).strip() if item[1] is not None else ""
        elif isinstance(item, dict):
            email = None
            password = None
            for k, v in item.items():
                key = str(k).strip().lower()
                if key in EMAIL_KEYS and email is None:
                    email = str(v).strip()
                elif key in PASSWORD_KEYS and password is None:
                    password = str(v).strip() if v is not None else ""
            if email is None:
                continue
        else:
            continue
        if email and "@" in email:
            accounts.append({"email": email, "password": password or ""})
    return accounts
